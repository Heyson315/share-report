from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Constants for sample data
INITIAL_BUDGET = 10000
OFFICE_SUPPLIES_EXPENSE = 500
REMAINING_BALANCE = 9500
OFFICE_SUPPLIES_BUDGET = 2000
MARKETING_BUDGET = 5000
DEVELOPMENT_BUDGET = 3000


def create_project_management_workbook(filename=None):
    """
    Create a project management Excel workbook.

    Args:
        filename (str, optional): Output filename for the workbook. Defaults to 'Project_Management.xlsx'.
    """
    if filename is None:
        filename = "Project_Management.xlsx"
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Financial Transactions Sheet
    transactions_sheet = workbook.active
    transactions_sheet.title = "Financial Transactions"
    trans_headers = ["Date", "Description", "Category", "Income", "Expense", "Balance"]
    for column_index, header in enumerate(trans_headers, 1):
        cell = transactions_sheet.cell(row=1, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Project Tasks Sheet
    tasks_sheet = workbook.create_sheet(title="Project Tasks")
    task_headers = ["Task ID", "Task Name", "Start Date", "Due Date", "Status", "Assigned To", "Notes"]
    for column_index, header in enumerate(task_headers, 1):
        cell = tasks_sheet.cell(row=1, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

    # Budget Summary Sheet
    budget_sheet = workbook.create_sheet(title="Budget Summary")
    budget_headers = ["Category", "Budgeted", "Spent", "Remaining", "Percent Spent"]
    for column_index, header in enumerate(budget_headers, 1):
        cell = budget_sheet.cell(row=1, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")

    # Adjust column widths for all sheets
    for sheet in workbook.worksheets:
        for column_index in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column_index)].width = 15

    # Sample transactions
    sample_transactions = [
        [datetime.now().strftime("%Y-%m-%d"), "Initial Budget", "Budget", INITIAL_BUDGET, 0, INITIAL_BUDGET],
        [
            datetime.now().strftime("%Y-%m-%d"),
            "Office Supplies",
            "Expenses",
            0,
            OFFICE_SUPPLIES_EXPENSE,
            REMAINING_BALANCE,
        ],
    ]

    # Add sample data
    # Transactions
    for row_number, transaction_data in enumerate(sample_transactions, 2):
        for column_index, value in enumerate(transaction_data, 1):
            transactions_sheet.cell(row=row_number, column=column_index).value = value

    # Tasks
    sample_tasks = [
        [
            1,
            "Design",
            datetime.now().strftime("%Y-%m-%d"),
            datetime.now().strftime("%Y-%m-%d"),
            "In Progress",
            "Alice",
            "",
        ],
        [
            2,
            "Development",
            datetime.now().strftime("%Y-%m-%d"),
            datetime.now().strftime("%Y-%m-%d"),
            "Not Started",
            "Bob",
            "",
        ],
    ]
    for row_number, task_data in enumerate(sample_tasks, 2):
        for column_index, value in enumerate(task_data, 1):
            tasks_sheet.cell(row=row_number, column=column_index).value = value

    # Budget Categories
    sample_budget = [
        ["Office Supplies", OFFICE_SUPPLIES_BUDGET, OFFICE_SUPPLIES_EXPENSE, "=B2-C2", "=C2/B2*100"],
        ["Marketing", MARKETING_BUDGET, 0, "=B3-C3", "=C3/B3*100"],
        ["Development", DEVELOPMENT_BUDGET, 0, "=B4-C4", "=C4/B4*100"],
    ]
    for row_number, budget_data in enumerate(sample_budget, 2):
        for column_index, value in enumerate(budget_data, 1):
            budget_sheet.cell(row=row_number, column=column_index).value = value

    # Save the workbook
    workbook.save(filename)
    return workbook


# Example usage:
# create_project_management_workbook('Project_Management.xlsx')
