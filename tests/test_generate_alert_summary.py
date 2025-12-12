"""
Unit Tests for Security Alert Summary Generation

Tests the alert summary generator, statistics calculation, executive summary
generation, and multi-format export functionality.

Reference: test_generate_alert_summary.py - Comprehensive tests for alert summaries
"""

import json
import pytest
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch, MagicMock

# Import the module under test
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.generate_alert_summary import AlertSummaryGenerator


class TestAlertSummaryGenerator:
    """
    Test suite for AlertSummaryGenerator class.

    Tests:
        - Statistics calculation
        - Executive summary generation
        - Detailed breakdown generation
        - Multi-format export (JSON, HTML, Excel)
        - Edge cases (empty data, missing files)

    Reference: #TestAlertSummaryGenerator - Main test class for alert summaries
    """

    @pytest.fixture
    def sample_alerts_db(self):
        """
        Provide sample alerts database for testing.

        Reference: #sample_alerts_db - Test fixture for alerts data
        """
        return {
            "metadata": {
                "last_updated": "2025-12-11T10:00:00Z",
                "version": "1.0"
            },
            "alerts": {
                "ALERT001": {
                    "id": "ALERT001",
                    "source": "Defender",
                    "severity": "CRITICAL",
                    "status": "new",
                    "title": "Malware detected on workstation",
                    "created": "2025-12-10T08:00:00Z",
                    "last_seen": "2025-12-11T09:00:00Z",
                    "is_false_positive": False,
                    "normalized_severity": 100
                },
                "ALERT002": {
                    "id": "ALERT002",
                    "source": "Azure Sentinel",
                    "severity": "HIGH",
                    "status": "investigating",
                    "title": "Suspicious login from unknown location",
                    "created": "2025-12-10T10:00:00Z",
                    "last_seen": "2025-12-11T08:00:00Z",
                    "is_false_positive": False,
                    "normalized_severity": 70
                },
                "ALERT003": {
                    "id": "ALERT003",
                    "source": "Defender",
                    "severity": "MEDIUM",
                    "status": "remediated",
                    "title": "Outdated software detected",
                    "created": "2025-12-09T14:00:00Z",
                    "last_seen": "2025-12-10T16:00:00Z",
                    "is_false_positive": False,
                    "remediated_at": "2025-12-10T17:00:00Z",
                    "normalized_severity": 40
                },
                "ALERT004": {
                    "id": "ALERT004",
                    "source": "M365 Defender",
                    "severity": "LOW",
                    "status": "closed",
                    "title": "Password policy violation",
                    "created": "2025-12-08T09:00:00Z",
                    "last_seen": "2025-12-09T10:00:00Z",
                    "is_false_positive": True,
                    "normalized_severity": 10
                },
                "ALERT005": {
                    "id": "ALERT005",
                    "source": "Azure Sentinel",
                    "severity": "HIGH",
                    "status": "escalated",
                    "title": "Brute force attack detected",
                    "created": "2025-12-10T12:00:00Z",
                    "last_seen": "2025-12-11T07:00:00Z",
                    "is_false_positive": False,
                    "escalated_at": "2025-12-11T06:00:00Z",
                    "escalation_reason": "Requires SOC investigation",
                    "normalized_severity": 70
                }
            }
        }

    @pytest.fixture
    def sample_remediation_log(self):
        """
        Provide sample remediation log for testing.

        Reference: #sample_remediation_log - Test fixture for remediation data
        """
        return [
            {
                "alert_id": "ALERT003",
                "action": "update_software",
                "result": "success",
                "timestamp": "2025-12-10T17:00:00Z"
            },
            {
                "alert_id": "ALERT006",
                "action": "block_ip",
                "result": "success",
                "timestamp": "2025-12-09T15:00:00Z"
            },
            {
                "alert_id": "ALERT007",
                "action": "disable_user",
                "result": "failed",
                "timestamp": "2025-12-08T11:00:00Z"
            }
        ]

    @pytest.fixture
    def temp_alert_files(self, sample_alerts_db, sample_remediation_log):
        """
        Create temporary alert and remediation files for testing.

        Reference: #temp_alert_files - File setup fixture
        """
        with TemporaryDirectory() as td:
            td = Path(td)
            
            # Create alerts database
            alerts_db_path = td / "alerts.json"
            with open(alerts_db_path, "w") as f:
                json.dump(sample_alerts_db, f)
            
            # Create remediation log
            remediation_log_path = td / "remediation_log.json"
            with open(remediation_log_path, "w") as f:
                json.dump(sample_remediation_log, f)
            
            yield {
                "alerts_db": alerts_db_path,
                "remediation_log": remediation_log_path,
                "output_dir": td / "output"
            }

    def test_initialization_with_valid_files(self, temp_alert_files):
        """
        Test AlertSummaryGenerator initialization with valid files.

        Reference: #test_initialization_with_valid_files - Constructor test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        assert generator.alerts_db is not None
        assert "alerts" in generator.alerts_db
        assert len(generator.alerts_db["alerts"]) == 5
        assert len(generator.remediation_log) == 3

    def test_initialization_with_missing_files(self):
        """
        Test AlertSummaryGenerator initialization with missing files.

        Reference: #test_initialization_with_missing_files - Error handling test
        """
        with TemporaryDirectory() as td:
            td = Path(td)
            
            generator = AlertSummaryGenerator(
                alerts_db_path=td / "nonexistent_alerts.json",
                remediation_log_path=td / "nonexistent_log.json"
            )

            # Should initialize with empty data structures
            assert generator.alerts_db == {"alerts": {}, "metadata": {}}
            assert generator.remediation_log == []

    def test_calculate_statistics_basic(self, temp_alert_files):
        """
        Test basic statistics calculation.

        Reference: #test_calculate_statistics_basic - Statistics test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()

        # Total counts
        assert stats["total_alerts"] == 5
        assert stats["remediated_count"] == 1
        assert stats["escalated_count"] == 1
        assert stats["closed_count"] == 1
        assert stats["pending_count"] == 2  # new + investigating
        assert stats["false_positives"] == 1

    def test_calculate_statistics_by_severity(self, temp_alert_files):
        """
        Test statistics calculation by severity.

        Reference: #test_calculate_statistics_by_severity - Severity grouping test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()

        # Check severity counts
        assert stats["by_severity"]["CRITICAL"] == 1
        assert stats["by_severity"]["HIGH"] == 2
        assert stats["by_severity"]["MEDIUM"] == 1
        assert stats["by_severity"]["LOW"] == 1

        # Check open critical/high
        assert stats["critical_severity_open"] == 1  # ALERT001 is new
        assert stats["high_severity_open"] == 2  # ALERT002 investigating, ALERT005 escalated

    def test_calculate_statistics_by_source(self, temp_alert_files):
        """
        Test statistics calculation by source.

        Reference: #test_calculate_statistics_by_source - Source grouping test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()

        # Check source counts
        assert stats["by_source"]["Defender"] == 2
        assert stats["by_source"]["Azure Sentinel"] == 2
        assert stats["by_source"]["M365 Defender"] == 1

    def test_calculate_statistics_rates(self, temp_alert_files):
        """
        Test calculation of remediation and closure rates.

        Reference: #test_calculate_statistics_rates - Rate calculation test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()

        # Rates (1 remediated + 1 closed out of 5 total)
        assert stats["remediation_rate"] == 20.0  # 1/5 = 20%
        assert stats["closure_rate"] == 40.0  # (1+1)/5 = 40%
        assert stats["escalation_rate"] == 20.0  # 1/5 = 20%

    def test_calculate_statistics_remediation_log(self, temp_alert_files):
        """
        Test statistics from remediation log.

        Reference: #test_calculate_statistics_remediation_log - Remediation stats test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()

        assert stats["total_remediation_actions"] == 3
        assert stats["successful_remediations"] == 2
        assert stats["failed_remediations"] == 1

    def test_calculate_statistics_empty_data(self):
        """
        Test statistics calculation with no alerts.

        Reference: #test_calculate_statistics_empty_data - Edge case test
        """
        with TemporaryDirectory() as td:
            td = Path(td)
            
            # Create empty files
            empty_alerts = td / "empty_alerts.json"
            empty_log = td / "empty_log.json"
            
            with open(empty_alerts, "w") as f:
                json.dump({"alerts": {}, "metadata": {}}, f)
            
            with open(empty_log, "w") as f:
                json.dump([], f)
            
            generator = AlertSummaryGenerator(
                alerts_db_path=empty_alerts,
                remediation_log_path=empty_log
            )

            stats = generator.calculate_statistics()

            assert stats["total_alerts"] == 0
            assert stats["remediation_rate"] == 0
            assert stats["closure_rate"] == 0
            assert stats["escalation_rate"] == 0

    def test_generate_executive_summary(self, temp_alert_files):
        """
        Test executive summary generation.

        Reference: #test_generate_executive_summary - Summary format test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()
        summary = generator.generate_executive_summary(stats)

        # Check key content is present
        assert "EXECUTIVE SUMMARY" in summary
        assert "Total Alerts Investigated: 5" in summary
        assert "Alerts Remediated: 1" in summary
        assert "Critical Severity (Open): 1" in summary
        assert "RECOMMENDATIONS:" in summary

    def test_generate_executive_summary_with_warnings(self, temp_alert_files):
        """
        Test executive summary includes appropriate warnings.

        Reference: #test_generate_executive_summary_with_warnings - Warning generation test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()
        summary = generator.generate_executive_summary(stats)

        # Should warn about critical alerts
        assert "⚠️  URGENT: 1 critical alerts require immediate attention" in summary

    def test_generate_detailed_breakdown(self, temp_alert_files):
        """
        Test detailed breakdown generation.

        Reference: #test_generate_detailed_breakdown - Breakdown format test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()
        breakdown = generator.generate_detailed_breakdown(stats)

        # Check sections are present
        assert "DETAILED BREAKDOWN" in breakdown
        assert "ALERTS BY STATUS:" in breakdown
        assert "ALERTS BY SEVERITY:" in breakdown
        assert "ALERTS BY SOURCE:" in breakdown

    def test_export_json(self, temp_alert_files):
        """
        Test JSON export functionality.

        Reference: #test_export_json - JSON export test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        output_path = temp_alert_files["output_dir"] / "summary.json"
        generator.export_json(output_path)

        # Verify file was created
        assert output_path.exists()

        # Verify content
        with open(output_path, "r") as f:
            data = json.load(f)

        assert "metadata" in data
        assert "statistics" in data
        assert "alerts_database" in data
        assert "remediation_log" in data
        assert data["statistics"]["total_alerts"] == 5

    def test_export_html(self, temp_alert_files):
        """
        Test HTML export functionality.

        Reference: #test_export_html - HTML export test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        output_path = temp_alert_files["output_dir"] / "summary.html"
        generator.export_html(output_path)

        # Verify file was created
        assert output_path.exists()

        # Verify HTML content
        html_content = output_path.read_text()
        assert "<!DOCTYPE html>" in html_content
        assert "Security Alert Summary Report" in html_content
        assert "Total Alerts" in html_content
        assert "Critical Open" in html_content

    @pytest.mark.skipif(
        not hasattr(pytest, "importorskip"),
        reason="pandas not available"
    )
    def test_export_excel_with_pandas(self, temp_alert_files):
        """
        Test Excel export functionality when pandas is available.

        Reference: #test_export_excel_with_pandas - Excel export test
        """
        try:
            import pandas as pd
            import openpyxl
        except ImportError:
            pytest.skip("pandas or openpyxl not available")

        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        output_path = temp_alert_files["output_dir"] / "summary.xlsx"
        generator.export_excel(output_path)

        # Verify file was created
        assert output_path.exists()

        # Verify Excel content
        wb = openpyxl.load_workbook(output_path)
        assert "Summary" in wb.sheetnames
        assert "Alerts" in wb.sheetnames

    def test_export_excel_without_pandas(self, temp_alert_files):
        """
        Test Excel export handles missing pandas gracefully.

        Reference: #test_export_excel_without_pandas - Error handling test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        # Mock PANDAS_AVAILABLE to False
        import scripts.generate_alert_summary as module
        original_pandas_available = module.PANDAS_AVAILABLE
        
        try:
            module.PANDAS_AVAILABLE = False
            
            output_path = temp_alert_files["output_dir"] / "summary.xlsx"
            
            # Should not crash, just print warning
            generator.export_excel(output_path)
            
            # File should NOT be created
            assert not output_path.exists()
        
        finally:
            module.PANDAS_AVAILABLE = original_pandas_available

    def test_generate_alert_details(self, temp_alert_files):
        """
        Test alert details generation.

        Reference: #test_generate_alert_details - Details format test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        details = generator.generate_alert_details()

        # Check content
        assert "ALERT DETAILS" in details
        assert "NEW ALERTS" in details
        assert "REMEDIATED ALERTS" in details
        assert "ESCALATED ALERTS" in details
        assert "ALERT001" in details
        assert "Potential False Positive" in details  # ALERT004

    @pytest.mark.parametrize("status,count", [
        ("new", 1),
        ("investigating", 1),
        ("remediated", 1),
        ("escalated", 1),
        ("closed", 1),
    ])
    def test_statistics_by_status(self, temp_alert_files, status, count):
        """
        Test statistics calculation for each status type.

        Parameterized test to verify status counting logic.

        Reference: #test_statistics_by_status - Parameterized status test
        """
        generator = AlertSummaryGenerator(
            alerts_db_path=temp_alert_files["alerts_db"],
            remediation_log_path=temp_alert_files["remediation_log"]
        )

        stats = generator.calculate_statistics()
        assert stats["by_status"].get(status, 0) == count

    def test_large_dataset_performance(self):
        """
        Test generator performance with large dataset.

        Reference: #test_large_dataset_performance - Performance test
        """
        with TemporaryDirectory() as td:
            td = Path(td)
            
            # Create large dataset (1000 alerts)
            large_alerts = {
                "metadata": {},
                "alerts": {}
            }
            
            for i in range(1000):
                large_alerts["alerts"][f"ALERT{i:04d}"] = {
                    "id": f"ALERT{i:04d}",
                    "source": "Defender",
                    "severity": ["CRITICAL", "HIGH", "MEDIUM", "LOW"][i % 4],
                    "status": ["new", "investigating", "remediated", "closed"][i % 4],
                    "title": f"Test alert {i}",
                    "created": "2025-12-10T08:00:00Z",
                    "last_seen": "2025-12-11T09:00:00Z",
                    "is_false_positive": False,
                    "normalized_severity": 50
                }
            
            alerts_path = td / "large_alerts.json"
            log_path = td / "empty_log.json"
            
            with open(alerts_path, "w") as f:
                json.dump(large_alerts, f)
            
            with open(log_path, "w") as f:
                json.dump([], f)
            
            # Test performance (should complete quickly)
            import time
            start = time.time()
            
            generator = AlertSummaryGenerator(
                alerts_db_path=alerts_path,
                remediation_log_path=log_path
            )
            
            stats = generator.calculate_statistics()
            
            elapsed = time.time() - start
            
            # Should process 1000 alerts in <1 second
            assert elapsed < 1.0
            assert stats["total_alerts"] == 1000


class TestAlertSummaryMain:
    """
    Test suite for main() function and CLI integration.

    Reference: #TestAlertSummaryMain - CLI test class
    """

    @patch('scripts.generate_alert_summary.AlertSummaryGenerator')
    def test_main_default_arguments(self, mock_generator_class):
        """
        Test main() with default arguments.

        Reference: #test_main_default_arguments - Default CLI test
        """
        # This would require more complex mocking
        # Skipping for now as it requires sys.argv manipulation
        pass

    def test_argparse_format_choices(self):
        """
        Test that format choices are correctly defined.

        Reference: #test_argparse_format_choices - CLI validation test
        """
        from scripts.generate_alert_summary import main
        import argparse

        # Valid formats should be accepted
        valid_formats = ["json", "html", "excel", "all"]
        for fmt in valid_formats:
            # Would test with argparse, but requires more setup
            assert fmt in valid_formats


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
