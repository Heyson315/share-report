"""
Unit tests for the M365 CIS report generation script.
"""

import json
import os
from pathlib import Path
from tempfile import TemporaryDirectory
import pandas as pd
from openpyxl import load_workbook

from scripts.m365_cis_report import (
    load_cis_data,
    calculate_statistics,
    generate_excel_report,
    main as m365_cis_report_main,
)

# Sample CIS data for testing
SAMPLE_CIS_DATA = {
    "Timestamp": "2025-11-14T10:00:00Z",
    "Controls": [
        {
            "ControlId": "1.1",
            "Title": "Test Control 1",
            "Severity": "High",
            "Status": "Pass",
        },
        {
            "ControlId": "1.2",
            "Title": "Test Control 2",
            "Severity": "Medium",
            "Status": "Fail",
        },
        {
            "ControlId": "1.3",
            "Title": "Test Control 3",
            "Severity": "Low",
            "Status": "Manual",
        },
    ],
}


def test_load_cis_data():
    """Test loading of CIS data from a JSON file."""
    with TemporaryDirectory() as td:
        td = Path(td)
        json_file = td / "cis.json"
        json_file.write_text(json.dumps(SAMPLE_CIS_DATA), encoding="utf-8")

        data = load_cis_data(str(json_file))
        assert data is not None
        assert "Controls" in data
        assert len(data["Controls"]) == 3


def test_calculate_statistics():
    """Test calculation of statistics from CIS data."""
    stats = calculate_statistics(SAMPLE_CIS_DATA["Controls"])
    assert stats["total_controls"] == 3
    assert stats["passed"] == 1
    assert stats["failed"] == 1
    assert stats["manual"] == 1
    assert stats["pass_rate"] == "33.33%"


def test_generate_excel_report():
    """Test generation of the Excel report."""
    with TemporaryDirectory() as td:
        td = Path(td)
        output_file = td / "report.xlsx"
        stats = calculate_statistics(SAMPLE_CIS_DATA["Controls"])

        generate_excel_report(
            SAMPLE_CIS_DATA["Controls"], stats, str(output_file)
        )

        assert output_file.exists()

        # Validate workbook content
        wb = load_workbook(output_file)
        assert "Summary" in wb.sheetnames
        assert "Controls" in wb.sheetnames

        # Check summary sheet
        summary_ws = wb["Summary"]
        assert summary_ws["B2"].value == 3  # Total Controls
        assert summary_ws["B3"].value == 1  # Passed
        assert summary_ws["B4"].value == 1  # Failed

        # Check controls sheet
        controls_ws = wb["Controls"]
        assert controls_ws.max_row == 4  # Header + 3 controls
        assert controls_ws["A2"].value == "1.1"
        assert controls_ws["D2"].value == "Pass"


def test_main_script_execution(monkeypatch):
    """Test the main execution flow of the script."""
    with TemporaryDirectory() as td:
        td = Path(td)
        input_dir = td / "input"
        output_dir = td / "output"
        input_dir.mkdir()
        output_dir.mkdir()

        # Create a dummy input file
        input_file = input_dir / "m365_cis_audit_latest.json"
        input_file.write_text(json.dumps(SAMPLE_CIS_DATA), encoding="utf-8")

        # Mock the paths
        monkeypatch.setattr(
            "scripts.m365_cis_report.INPUT_DIR", str(input_dir)
        )
        monkeypatch.setattr(
            "scripts.m365_cis_report.OUTPUT_DIR", str(output_dir)
        )

        # Run the main function
        m365_cis_report_main()

        # Check for output file
        output_files = list(output_dir.glob("M365_CIS_Report_*.xlsx"))
        assert len(output_files) == 1
        assert output_files[0].exists()
