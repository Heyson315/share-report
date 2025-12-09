"""
Tests for the Excel report generation.
"""

from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from src.core.excel_generator import create_project_management_workbook


class TestCreateProjectManagementWorkbook:
    """Tests for the create_project_management_workbook function."""

    def test_creates_workbook_with_default_name(self):
        """Test that a workbook is created with the default filename."""
        with TemporaryDirectory() as td:
            # Change working directory to temp dir to avoid cluttering project root
            import os

            original_cwd = os.getcwd()
            os.chdir(td)

            create_project_management_workbook()

            report_path = Path("Project_Management.xlsx")
            assert report_path.exists()

            os.chdir(original_cwd)

    def test_creates_workbook_with_custom_name(self):
        """Test that a workbook is created with a custom filename."""
        with TemporaryDirectory() as td:
            report_path = Path(td) / "custom_name.xlsx"
            create_project_management_workbook(filename=str(report_path))
            assert report_path.exists()

    def test_workbook_has_correct_sheets_and_headers(self):
        """Test that the workbook contains the correct sheets and headers."""
        with TemporaryDirectory() as td:
            report_path = Path(td) / "test.xlsx"
            create_project_management_workbook(filename=str(report_path))

            workbook = openpyxl.load_workbook(report_path)

            # Check sheet names
            assert "Financial Transactions" in workbook.sheetnames
            assert "Project Tasks" in workbook.sheetnames
            assert "Budget Summary" in workbook.sheetnames

            # Check headers for Financial Transactions sheet
            transactions_sheet = workbook["Financial Transactions"]
            trans_headers = ["Date", "Description", "Category", "Income", "Expense", "Balance"]
            for col, header in enumerate(trans_headers, 1):
                assert transactions_sheet.cell(row=1, column=col).value == header

            # Check headers for Project Tasks sheet
            tasks_sheet = workbook["Project Tasks"]
            task_headers = ["Task ID", "Task Name", "Start Date", "Due Date", "Status", "Assigned To", "Notes"]
            for col, header in enumerate(task_headers, 1):
                assert tasks_sheet.cell(row=1, column=col).value == header

            # Check headers for Budget Summary sheet
            budget_sheet = workbook["Budget Summary"]
            budget_headers = ["Category", "Budgeted", "Spent", "Remaining", "Percent Spent"]
            for col, header in enumerate(budget_headers, 1):
                assert budget_sheet.cell(row=1, column=col).value == header

    def test_workbook_contains_sample_data(self):
        """Test that the workbook is populated with sample data."""
        with TemporaryDirectory() as td:
            report_path = Path(td) / "test.xlsx"
            create_project_management_workbook(filename=str(report_path))

            workbook = openpyxl.load_workbook(report_path)

            # Check sample transaction data
            transactions_sheet = workbook["Financial Transactions"]
            assert transactions_sheet.cell(row=2, column=2).value == "Initial Budget"
            assert transactions_sheet.cell(row=3, column=5).value == 500

            # Check sample task data
            tasks_sheet = workbook["Project Tasks"]
            assert tasks_sheet.cell(row=2, column=2).value == "Design"
            assert tasks_sheet.cell(row=3, column=5).value == "Not Started"
