from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl
import pandas as pd
import pytest

from src.core.excel_generator import create_project_management_workbook
from src.integrations.sharepoint_connector import build_summaries, main, write_excel_report


@pytest.fixture
def sample_dataframe():
    """A fixture that provides a sample DataFrame for testing."""
    data = {
        "Resource Path": ["/sites/SiteA/doc1", "/sites/SiteA/doc2", "/sites/SiteB/doc3", "/sites/SiteA/doc1"],
        "Item Type": ["File", "File", "Folder", "File"],
        "Permission": ["Read", "Write", "Read", "Read"],
        "User Name": ["User1", "User2", "User1", "User3"],
        "User Email": ["user1@test.com", "user2@test.com", "user1@test.com", "user3@test.com"],
        "User Or Group Type": ["User", "User", "User", "User"],
        "Link ID": ["", "", "", ""],
        "Link Type": ["", "", "", ""],
        "AccessViaLinkID": ["", "", "", ""],
    }
    return pd.DataFrame(data)


class TestBuildSummaries:
    """Tests for the build_summaries function."""

    def test_build_summaries_with_basic_data(self, sample_dataframe):
        """Test that summaries are built correctly with a basic DataFrame."""
        summaries = build_summaries(sample_dataframe)
        assert "by_item_type" in summaries
        assert "by_permission" in summaries
        assert "top_users" in summaries
        assert "top_resources" in summaries

        assert summaries["by_item_type"].iloc[0]["Item Type"] == "File"
        assert summaries["by_item_type"].iloc[0]["Count"] == 3

        assert summaries["by_permission"].iloc[0]["Permission"] == "Read"
        assert summaries["by_permission"].iloc[0]["Count"] == 3

        assert summaries["top_users"].iloc[0]["User Email"] == "user1@test.com"
        assert summaries["top_users"].iloc[0]["Count"] == 2

        assert summaries["top_resources"].iloc[0]["Resource Path"] == "/sites/SiteA/doc1"
        assert summaries["top_resources"].iloc[0]["Count"] == 2

    def test_build_summaries_with_missing_columns(self):
        """Test that summaries are built correctly when some columns are missing."""
        data = {
            "Resource Path": ["/sites/SiteA/doc1"],
            "Permission": ["Read"],
        }
        df = pd.DataFrame(data)
        summaries = build_summaries(df)
        assert "by_item_type" not in summaries
        assert "by_permission" in summaries
        assert "top_users" not in summaries
        assert "top_resources" in summaries

    def test_build_summaries_with_empty_strings(self):
        """Test that summaries handle empty strings gracefully."""
        data = {
            "Resource Path": ["/sites/SiteA/doc1", ""],
            "Item Type": ["File", "File"],
            "Permission": ["Read", "Read"],
            "User Name": ["User1", ""],
            "User Email": ["user1@test.com", ""],
        }
        df = pd.DataFrame(data)
        summaries = build_summaries(df)
        assert summaries["top_users"].shape[0] == 1
        assert summaries["top_resources"].shape[0] == 1

    def test_build_summaries_normalization(self):
        """Test that string columns are normalized (stripped and converted)."""
        data = {
            "Item Type": [" File ", "Folder"],  # Whitespace
            "Permission": [123, " Read "],  # Non-string and whitespace
        }
        df = pd.DataFrame(data)
        summaries = build_summaries(df)

        # Check Item Type normalization
        item_types = summaries["by_item_type"]
        assert "File" in item_types["Item Type"].values
        assert " File " not in item_types["Item Type"].values

        # Check Permission normalization
        permissions = summaries["by_permission"]
        assert "123" in permissions["Permission"].values
        assert "Read" in permissions["Permission"].values


class TestWriteExcelReport:
    """Tests for the write_excel_report function."""

    def test_write_excel_report_creates_file(self, sample_dataframe):
        """Test that an Excel report is created."""
        summaries = build_summaries(sample_dataframe)
        with TemporaryDirectory() as td:
            output_path = Path(td) / "report.xlsx"
            write_excel_report(summaries, output_path)
            assert output_path.exists()

    def test_write_excel_report_has_correct_sheets(self, sample_dataframe):
        """Test that the created Excel report has the correct sheets."""
        summaries = build_summaries(sample_dataframe)
        with TemporaryDirectory() as td:
            output_path = Path(td) / "report.xlsx"
            write_excel_report(summaries, output_path)

            workbook = openpyxl.load_workbook(output_path)
            assert "Overview" in workbook.sheetnames
            assert "by_item_type" in workbook.sheetnames
            assert "by_permission" in workbook.sheetnames
            assert "top_users" in workbook.sheetnames
            assert "top_resources" in workbook.sheetnames

    def test_write_excel_report_overview_content(self, sample_dataframe):
        """Test that the Overview sheet contains correct summary statistics."""
        summaries = build_summaries(sample_dataframe)
        with TemporaryDirectory() as td:
            output_path = Path(td) / "report.xlsx"
            write_excel_report(summaries, output_path)

            # Read back the Overview sheet
            overview_df = pd.read_excel(output_path, sheet_name="Overview")

            # Check columns
            assert "Summary" in overview_df.columns
            assert "Rows" in overview_df.columns
            assert "Columns" in overview_df.columns

            # Check that we have rows for each summary
            assert len(overview_df) == len(summaries)
            assert "by_item_type" in overview_df["Summary"].values


class TestMainFunction:
    """Tests for the main function."""

    def test_main_with_args(self, monkeypatch, sample_dataframe):
        """Test the main function with command-line arguments."""
        with TemporaryDirectory() as td:
            td_path = Path(td)
            input_file = td_path / "input.csv"
            output_file = td_path / "output.xlsx"
            sample_dataframe.to_csv(input_file, index=False)

            monkeypatch.setattr(
                "sys.argv",
                ["", f"--input={input_file}", f"--output={output_file}"],
            )
            main()
            assert output_file.exists()

    def test_create_project_management_workbook(self):
        """Test that the create_project_management_workbook function runs without error."""
        with TemporaryDirectory() as td:
            output_path = Path(td) / "workbook.xlsx"
            create_project_management_workbook(str(output_path))
            assert output_path.exists()
