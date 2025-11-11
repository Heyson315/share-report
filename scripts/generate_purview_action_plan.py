"""Generate Purview Audit Retention Action Plan Excel Report"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_purview_action_plan():
    """Create Excel workbook with Purview audit retention action plan"""

    # Create output directory
    output_dir = Path("output/reports/business")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # --- Sheet 1: Executive Summary ---
    ws_summary = wb.create_sheet("Executive Summary")

    summary_data = [
        ["Purview Audit Retention - Action Plan"],
        ["Rahman Finance and Accounting P.L.LC"],
        [""],
        ["Report Date", datetime.now().strftime("%B %d, %Y")],
        ["Current Compliance Status", "Manual Review Required"],
        ["Target Compliance", "CIS M365 Foundations Benchmark v3.0 Level 1"],
        [""],
        ["Current State"],
        ["• E5 License: 1-year default audit retention (365 days)"],
        ["• CIS Requirement: 90+ days retention (ALREADY MET)"],
        ["• Mailbox Auditing: Enabled ✓"],
        ["• Unified Audit Log: Not explicitly configured"],
        [""],
        ["Recommended Actions"],
        ["1. Configure explicit audit retention policy (Portal method recommended)"],
        ["2. Extend retention to 2-3 years for CPA compliance"],
        ["3. Document policy for SOC 2 / IRS requirements"],
        ["4. Schedule quarterly retention policy reviews"],
        [""],
        ["Business Impact"],
        ["• IRS Compliance: Covers 3-year statute of limitations"],
        ["• Client Audits: Complete audit trail for client work"],
        ["• Fraud Detection: Historical investigation capability"],
        ["• SOC 2: Evidence for security controls"],
        ["• E-Discovery: Legal hold and investigation support"],
    ]

    for row_idx, row_data in enumerate(summary_data, start=1):
        if isinstance(row_data, list) and len(row_data) == 1:
            # Header rows
            cell = ws_summary.cell(row=row_idx, column=1, value=row_data[0])
            if row_idx == 1:
                cell.font = Font(bold=True, size=16, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            elif row_idx == 2:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            elif (
                "Current State" in row_data[0]
                or "Recommended Actions" in row_data[0]
                or "Business Impact" in row_data[0]
            ):
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        elif isinstance(row_data, list) and len(row_data) == 2:
            # Key-value rows
            ws_summary.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=row_data[1])

    # Merge cells for headers
    ws_summary.merge_cells("A1:B1")
    ws_summary.merge_cells("A2:B2")

    # Set column widths
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 60

    # --- Sheet 2: Implementation Steps ---
    ws_steps = wb.create_sheet("Implementation Steps")

    steps_data = [
        {
            "Step": 1,
            "Action": "Access Purview Compliance Portal",
            "Method": "Portal (Recommended)",
            "Instructions": "1. Navigate to https://compliance.microsoft.com\n2. Sign in with Hassan@hhr-cpa.us\n3. Go to Solutions > Audit > Audit retention policies",
            "Time Required": "5 minutes",
            "Prerequisites": "Global Admin or Compliance Admin role",
            "Status": "Not Started",
        },
        {
            "Step": 2,
            "Action": "Create Audit Retention Policy",
            "Method": "Portal",
            "Instructions": "1. Click '+ Create audit retention policy'\n2. Name: 'CPA Firm Audit Retention - 3 Years'\n3. Description: 'Retain audit logs for 3 years per IRS requirements'\n4. Duration: 1095 days (3 years)\n5. Record types: Select all or minimum (Exchange, SharePoint, OneDrive, Azure AD)\n6. Users: All users\n7. Priority: 1",
            "Time Required": "10 minutes",
            "Prerequisites": "Step 1 completed",
            "Status": "Not Started",
        },
        {
            "Step": 3,
            "Action": "Verify Policy Configuration",
            "Method": "Portal",
            "Instructions": "1. Return to Audit retention policies page\n2. Confirm policy appears in list\n3. Check Status = 'On'\n4. Verify Duration = 1095 days\n5. Note the Policy ID for documentation",
            "Time Required": "5 minutes",
            "Prerequisites": "Step 2 completed",
            "Status": "Not Started",
        },
        {
            "Step": 4,
            "Action": "Test Audit Log Search",
            "Method": "Portal",
            "Instructions": "1. Go to Solutions > Audit > Audit log search\n2. Search for recent activities (last 7 days)\n3. Verify results appear\n4. Confirm retention warning shows 3 years",
            "Time Required": "5 minutes",
            "Prerequisites": "Step 3 completed",
            "Status": "Not Started",
        },
        {
            "Step": 5,
            "Action": "Document Configuration",
            "Method": "Internal Documentation",
            "Instructions": "1. Screenshot the retention policy settings\n2. Update IT documentation with policy details\n3. Note configuration date and administrator\n4. Add to compliance documentation for SOC 2",
            "Time Required": "15 minutes",
            "Prerequisites": "Steps 1-4 completed",
            "Status": "Not Started",
        },
        {
            "Step": 6,
            "Action": "Re-run M365 CIS Audit",
            "Method": "PowerShell",
            "Instructions": "1. Open PowerShell in share-report directory\n2. Run: .\\scripts\\powershell\\Invoke-M365CISAudit.ps1 -Timestamped\n3. Verify CIS-PURVIEW-2 now shows 'Pass' or updated evidence\n4. Generate Excel report with: python scripts/m365_cis_report.py",
            "Time Required": "10 minutes",
            "Prerequisites": "Step 5 completed",
            "Status": "Not Started",
        },
    ]

    # Write headers
    headers = ["Step", "Action", "Method", "Instructions", "Time Required", "Prerequisites", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_steps.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, step in enumerate(steps_data, start=2):
        ws_steps.cell(row=row_idx, column=1, value=step["Step"]).alignment = Alignment(horizontal="center")
        ws_steps.cell(row=row_idx, column=2, value=step["Action"]).alignment = Alignment(wrap_text=True)
        ws_steps.cell(row=row_idx, column=3, value=step["Method"]).alignment = Alignment(horizontal="center")
        ws_steps.cell(row=row_idx, column=4, value=step["Instructions"]).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws_steps.cell(row=row_idx, column=5, value=step["Time Required"]).alignment = Alignment(horizontal="center")
        ws_steps.cell(row=row_idx, column=6, value=step["Prerequisites"]).alignment = Alignment(wrap_text=True)

        status_cell = ws_steps.cell(row=row_idx, column=7, value=step["Status"])
        status_cell.alignment = Alignment(horizontal="center")
        if step["Status"] == "Not Started":
            status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Set column widths
    ws_steps.column_dimensions["A"].width = 8
    ws_steps.column_dimensions["B"].width = 35
    ws_steps.column_dimensions["C"].width = 18
    ws_steps.column_dimensions["D"].width = 60
    ws_steps.column_dimensions["E"].width = 15
    ws_steps.column_dimensions["F"].width = 20
    ws_steps.column_dimensions["G"].width = 15

    # --- Sheet 3: PowerShell Alternative ---
    ws_ps = wb.create_sheet("PowerShell Method")

    ps_content = [
        ["PowerShell Configuration Method (Advanced)"],
        [""],
        ["Prerequisites:"],
        ["• ExchangeOnlineManagement module installed"],
        ["• Security & Compliance PowerShell connection"],
        ["• Compliance Administrator or Global Admin role"],
        [""],
        ["Step 1: Connect to Security & Compliance PowerShell"],
        [""],
        ["# Import module"],
        ["Import-Module ExchangeOnlineManagement"],
        [""],
        ["# Connect to Security & Compliance"],
        ["Connect-IPPSSession"],
        [""],
        ["Step 2: Create Audit Retention Policy"],
        [""],
        ["# Create 3-year retention policy"],
        ["New-UnifiedAuditLogRetentionPolicy `"],
        ["    -Name 'CPA-Audit-3Years' `"],
        ["    -Description '3-year retention for IRS compliance' `"],
        ["    -RetentionDuration TenYears `"],  # Note: Use predefined values
        ["    -RecordTypes @('ExchangeAdmin','ExchangeItem','SharePoint','OneDrive','AzureActiveDirectory') `"],
        ["    -Priority 1"],
        [""],
        ["Step 3: Verify Policy"],
        [""],
        ["# List all retention policies"],
        ["Get-UnifiedAuditLogRetentionPolicy"],
        [""],
        ["# Check specific policy"],
        ["Get-UnifiedAuditLogRetentionPolicy -Identity 'CPA-Audit-3Years' | Format-List"],
        [""],
        ["Step 4: View Audit Log (Test)"],
        [""],
        ["# Search audit logs from last 7 days"],
        ["Search-UnifiedAuditLog -StartDate (Get-Date).AddDays(-7) -EndDate (Get-Date) | Select-Object -First 10"],
        [""],
        ["Important Notes:"],
        [
            "• RetentionDuration accepts: ThreeDays, SevenDays, FourteenDays, OneMonth, ThreeMonths, SixMonths, NineMonths, TwelveMonths, TenYears"
        ],
        ["• For 3 years (1095 days), you may need to use custom duration or portal method"],
        ["• Policy changes take effect within 24 hours"],
        ["• Existing logs are retained according to new policy"],
    ]

    for row_idx, content in enumerate(ps_content, start=1):
        cell = ws_ps.cell(row=row_idx, column=1, value=content[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        elif "Step" in content[0] or "Prerequisites" in content[0] or "Important Notes" in content[0]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        elif content[0].startswith("#"):
            cell.font = Font(italic=True, color="008000")
        elif (
            content[0].startswith("New-")
            or content[0].startswith("Get-")
            or content[0].startswith("Connect-")
            or content[0].startswith("Search-")
            or content[0].startswith("Import-")
        ):
            cell.font = Font(name="Consolas", size=10)
        elif "    -" in content[0]:
            cell.font = Font(name="Consolas", size=9, color="4472C4")

    ws_ps.column_dimensions["A"].width = 100

    # --- Sheet 4: Compliance Checklist ---
    ws_checklist = wb.create_sheet("Compliance Checklist")

    checklist_data = [
        {
            "Requirement": "CIS M365 Foundations v3.0 L1",
            "Control": "CIS-PURVIEW-2",
            "Description": "Audit log retention ≥ 90 days",
            "Current Status": "COMPLIANT (E5 default: 365 days)",
            "Recommended": "Configure explicit policy for 1095 days (3 years)",
            "Priority": "Medium",
        },
        {
            "Requirement": "IRS Record Retention",
            "Control": "26 CFR § 1.6001-1",
            "Description": "Tax records retained for 3 years from filing",
            "Current Status": "NEEDS CONFIGURATION",
            "Recommended": "Set 3-year retention policy",
            "Priority": "High",
        },
        {
            "Requirement": "SOC 2 Compliance",
            "Control": "CC7.2 - Monitoring Activities",
            "Description": "System activities monitored and logged",
            "Current Status": "PARTIAL (logs exist but policy not documented)",
            "Recommended": "Document retention policy and procedures",
            "Priority": "High",
        },
        {
            "Requirement": "Client Audit Support",
            "Control": "Internal Policy",
            "Description": "Audit trail for client work and communications",
            "Current Status": "NEEDS IMPROVEMENT",
            "Recommended": "3-year retention for complete audit cycles",
            "Priority": "Medium",
        },
        {
            "Requirement": "Fraud Investigation",
            "Control": "Internal Control",
            "Description": "Historical data for investigating suspicious activity",
            "Current Status": "ADEQUATE (1 year default)",
            "Recommended": "3-year retention for thorough investigations",
            "Priority": "Medium",
        },
    ]

    # Write headers
    checklist_headers = ["Requirement", "Control", "Description", "Current Status", "Recommended", "Priority"]
    for col_idx, header in enumerate(checklist_headers, start=1):
        cell = ws_checklist.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data
    for row_idx, item in enumerate(checklist_data, start=2):
        ws_checklist.cell(row=row_idx, column=1, value=item["Requirement"]).alignment = Alignment(wrap_text=True)
        ws_checklist.cell(row=row_idx, column=2, value=item["Control"]).alignment = Alignment(wrap_text=True)
        ws_checklist.cell(row=row_idx, column=3, value=item["Description"]).alignment = Alignment(wrap_text=True)
        ws_checklist.cell(row=row_idx, column=4, value=item["Current Status"]).alignment = Alignment(wrap_text=True)
        ws_checklist.cell(row=row_idx, column=5, value=item["Recommended"]).alignment = Alignment(wrap_text=True)

        priority_cell = ws_checklist.cell(row=row_idx, column=6, value=item["Priority"])
        priority_cell.alignment = Alignment(horizontal="center")
        if item["Priority"] == "High":
            priority_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            priority_cell.font = Font(bold=True, color="FFFFFF")
        elif item["Priority"] == "Medium":
            priority_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Set column widths
    ws_checklist.column_dimensions["A"].width = 25
    ws_checklist.column_dimensions["B"].width = 20
    ws_checklist.column_dimensions["C"].width = 40
    ws_checklist.column_dimensions["D"].width = 35
    ws_checklist.column_dimensions["E"].width = 35
    ws_checklist.column_dimensions["F"].width = 12

    # --- Sheet 5: Quick Reference ---
    ws_ref = wb.create_sheet("Quick Reference")

    reference_data = [
        ["Purview Audit Retention - Quick Reference"],
        [""],
        ["Key URLs"],
        ["Purview Compliance Portal", "https://compliance.microsoft.com"],
        ["Audit Log Search", "https://compliance.microsoft.com/auditlogsearch"],
        ["Audit Retention Policies", "https://compliance.microsoft.com/auditlogretention"],
        ["Microsoft Documentation", "https://learn.microsoft.com/en-us/purview/audit-log-retention-policies"],
        [""],
        ["Current Configuration"],
        ["License", "Microsoft 365 E5 (no Teams)"],
        ["Default Retention", "365 days (1 year)"],
        ["Maximum Retention (E5)", "10 years"],
        ["Mailbox Auditing", "Enabled ✓"],
        ["Unified Audit Log", "Needs explicit policy configuration"],
        [""],
        ["Recommended Settings"],
        ["Policy Name", "CPA Firm Audit Retention - 3 Years"],
        ["Duration", "1095 days (3 years)"],
        ["Record Types", "All (or minimum: Exchange, SharePoint, OneDrive, Azure AD)"],
        ["Users", "All users"],
        ["Priority", "1"],
        [""],
        ["Retention Durations Available"],
        ["IRS Requirement", "3 years (26 CFR § 1.6001-1)"],
        ["CIS Minimum", "90 days (already met with E5 default)"],
        ["Recommended for CPA", "2-3 years (covers full audit cycles)"],
        ["SOC 2 Best Practice", "1-3 years depending on controls"],
        [""],
        ["Contact Information"],
        ["Administrator", "Hassan@hhr-cpa.us"],
        ["Tenant", "RahmanFinanceandAccounting.onmicrosoft.com"],
        ["Support", "Microsoft 365 Admin Center"],
    ]

    for row_idx, row_data in enumerate(reference_data, start=1):
        if isinstance(row_data, list):
            if len(row_data) == 1:
                cell = ws_ref.cell(row=row_idx, column=1, value=row_data[0])
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                    ws_ref.merge_cells(f"A{row_idx}:B{row_idx}")
                elif (
                    "Key URLs" in row_data[0]
                    or "Current Configuration" in row_data[0]
                    or "Recommended Settings" in row_data[0]
                    or "Retention Durations" in row_data[0]
                    or "Contact Information" in row_data[0]
                ):
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            elif len(row_data) == 2:
                ws_ref.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)
                cell_value = ws_ref.cell(row=row_idx, column=2, value=row_data[1])
                if row_data[1].startswith("http"):
                    cell_value.font = Font(color="0563C1", underline="single")

    ws_ref.column_dimensions["A"].width = 30
    ws_ref.column_dimensions["B"].width = 70

    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"purview_audit_retention_action_plan_{timestamp}.xlsx"
    wb.save(output_file)

    return output_file


if __name__ == "__main__":
    output_file = create_purview_action_plan()
    print(f"✅ Purview Action Plan created: {output_file}")
