#!/usr/bin/env python3
"""
generate_alert_summary.py

Generate comprehensive summary reports for security alert investigation and remediation.
Designed for compliance teams and security operations.

Features:
- Compliance-ready output format
- Detailed investigation and remediation statistics
- Trend analysis over time
- Executive summary with key metrics
- Export to multiple formats (JSON, HTML, Excel)

Usage:
    python scripts/generate_alert_summary.py
    python scripts/generate_alert_summary.py --format html
    python scripts/generate_alert_summary.py --format excel
"""

import argparse
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️  pandas not available - Excel export disabled")


# Color scheme constants for reports
HEADER_COLOR = "4472C4"  # Microsoft blue for headers


class AlertSummaryGenerator:
    """Generate comprehensive security alert summary reports."""

    def __init__(self, alerts_db_path: Path, remediation_log_path: Path):
        """
        Initialize the generator.

        Args:
            alerts_db_path: Path to alerts database
            remediation_log_path: Path to remediation log
        """
        self.alerts_db_path = alerts_db_path
        self.remediation_log_path = remediation_log_path
        self.alerts_db = self._load_alerts_db()
        self.remediation_log = self._load_remediation_log()

    def _load_alerts_db(self) -> Dict[str, Any]:
        """Load alerts database."""
        if not self.alerts_db_path.exists():
            print(f"⚠️  Alerts database not found: {self.alerts_db_path}")
            return {"alerts": {}, "metadata": {}}

        with open(self.alerts_db_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _load_remediation_log(self) -> List[Dict[str, Any]]:
        """Load remediation log."""
        if not self.remediation_log_path.exists():
            print(f"⚠️  Remediation log not found: {self.remediation_log_path}")
            return []

        with open(self.remediation_log_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def calculate_statistics(self) -> Dict[str, Any]:
        """
        Calculate comprehensive statistics.

        Returns:
            Statistics dictionary
        """
        stats = {
            "total_alerts": len(self.alerts_db["alerts"]),
            "by_status": {},
            "by_severity": {},
            "by_source": {},
            "false_positives": 0,
            "remediated_count": 0,
            "escalated_count": 0,
            "closed_count": 0,
            "pending_count": 0,
            "high_severity_open": 0,
            "critical_severity_open": 0,
            "total_remediation_actions": len(self.remediation_log),
            "successful_remediations": 0,
            "failed_remediations": 0,
        }

        # Analyze alerts
        for alert in self.alerts_db["alerts"].values():
            status = alert.get("status", "unknown")
            severity = alert.get("severity", "UNKNOWN").upper()
            source = alert.get("source", "unknown")

            # Count by status
            stats["by_status"][status] = stats["by_status"].get(status, 0) + 1

            # Count by severity
            stats["by_severity"][severity] = stats["by_severity"].get(severity, 0) + 1

            # Count by source
            stats["by_source"][source] = stats["by_source"].get(source, 0) + 1

            # Count false positives
            if alert.get("is_false_positive"):
                stats["false_positives"] += 1

            # Count by resolution
            if status == "remediated":
                stats["remediated_count"] += 1
            elif status == "escalated":
                stats["escalated_count"] += 1
            elif status == "closed":
                stats["closed_count"] += 1
            elif status in ["new", "investigating"]:
                stats["pending_count"] += 1

            # Count high/critical open alerts
            if status not in ["remediated", "closed"]:
                if severity == "HIGH":
                    stats["high_severity_open"] += 1
                elif severity == "CRITICAL":
                    stats["critical_severity_open"] += 1

        # Analyze remediation actions
        for action in self.remediation_log:
            result = action.get("result", "unknown")
            if result == "success":
                stats["successful_remediations"] += 1
            elif result in ["error", "failed"]:
                stats["failed_remediations"] += 1

        # Calculate rates
        if stats["total_alerts"] > 0:
            stats["remediation_rate"] = round((stats["remediated_count"] / stats["total_alerts"]) * 100, 2)
            stats["escalation_rate"] = round((stats["escalated_count"] / stats["total_alerts"]) * 100, 2)
            stats["closure_rate"] = round(
                ((stats["remediated_count"] + stats["closed_count"]) / stats["total_alerts"]) * 100, 2
            )
        else:
            stats["remediation_rate"] = 0
            stats["escalation_rate"] = 0
            stats["closure_rate"] = 0

        return stats

    def generate_executive_summary(self, stats: Dict[str, Any]) -> str:
        """
        Generate executive summary text.

        Args:
            stats: Statistics dictionary

        Returns:
            Executive summary as formatted string
        """
        summary = []
        summary.append("=" * 70)
        summary.append("EXECUTIVE SUMMARY - Security Alert Investigation & Remediation")
        summary.append("=" * 70)
        summary.append("")
        summary.append(f"Report Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}")
        summary.append("")
        summary.append("KEY METRICS:")
        summary.append(f"  • Total Alerts Investigated: {stats['total_alerts']}")
        summary.append(f"  • Alerts Remediated: {stats['remediated_count']} ({stats['remediation_rate']}%)")
        summary.append(f"  • Alerts Escalated: {stats['escalated_count']} ({stats['escalation_rate']}%)")
        summary.append(f"  • Alerts Closed: {stats['closed_count']}")
        summary.append(f"  • Pending Alerts: {stats['pending_count']}")
        summary.append("")
        summary.append("RISK ASSESSMENT:")
        summary.append(f"  • Critical Severity (Open): {stats['critical_severity_open']}")
        summary.append(f"  • High Severity (Open): {stats['high_severity_open']}")
        summary.append(f"  • False Positives Identified: {stats['false_positives']}")
        summary.append("")
        summary.append("REMEDIATION EFFECTIVENESS:")
        summary.append(f"  • Total Remediation Actions: {stats['total_remediation_actions']}")
        summary.append(f"  • Successful Remediations: {stats['successful_remediations']}")
        summary.append(f"  • Failed Remediations: {stats['failed_remediations']}")
        summary.append(f"  • Overall Closure Rate: {stats['closure_rate']}%")
        summary.append("")
        summary.append("RECOMMENDATIONS:")

        # Generate recommendations based on metrics
        if stats["critical_severity_open"] > 0:
            summary.append(
                f"  ⚠️  URGENT: {stats['critical_severity_open']} critical alerts require immediate attention"
            )

        if stats["high_severity_open"] > 5:
            summary.append(f"  ⚠️  {stats['high_severity_open']} high-severity alerts pending review")

        if stats["closure_rate"] < 50:
            summary.append("  ⚠️  Low closure rate - consider increasing remediation resources")

        if stats["escalated_count"] > stats["remediated_count"]:
            summary.append("  ⚠️  High escalation rate - review auto-remediation capabilities")

        if stats["false_positives"] > stats["total_alerts"] * 0.2:
            summary.append("  ⚠️  High false positive rate - tune detection rules")

        if not any("⚠️" in line for line in summary[-5:]):
            summary.append("  ✅ No critical issues identified")

        summary.append("")
        summary.append("=" * 70)

        return "\n".join(summary)

    def generate_detailed_breakdown(self, stats: Dict[str, Any]) -> str:
        """
        Generate detailed breakdown by status, severity, and source.

        Args:
            stats: Statistics dictionary

        Returns:
            Detailed breakdown as formatted string
        """
        breakdown = []
        breakdown.append("\nDETAILED BREAKDOWN")
        breakdown.append("=" * 70)

        breakdown.append("\nALERTS BY STATUS:")
        for status, count in sorted(stats["by_status"].items(), key=lambda x: -x[1]):
            percentage = (count / stats["total_alerts"]) * 100 if stats["total_alerts"] > 0 else 0
            breakdown.append(f"  {status:20} : {count:4} ({percentage:5.1f}%)")

        breakdown.append("\nALERTS BY SEVERITY:")
        severity_order = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO", "WARNING", "ERROR"]
        for severity in severity_order:
            if severity in stats["by_severity"]:
                count = stats["by_severity"][severity]
                percentage = (count / stats["total_alerts"]) * 100 if stats["total_alerts"] > 0 else 0
                breakdown.append(f"  {severity:20} : {count:4} ({percentage:5.1f}%)")

        breakdown.append("\nALERTS BY SOURCE:")
        for source, count in sorted(stats["by_source"].items(), key=lambda x: -x[1]):
            percentage = (count / stats["total_alerts"]) * 100 if stats["total_alerts"] > 0 else 0
            breakdown.append(f"  {source:20} : {count:4} ({percentage:5.1f}%)")

        breakdown.append("\n" + "=" * 70)

        return "\n".join(breakdown)

    def generate_alert_details(self) -> str:
        """
        Generate details for all alerts.

        Returns:
            Alert details as formatted string
        """
        details = []
        details.append("\nALERT DETAILS")
        details.append("=" * 70)

        # Group alerts by status
        alerts_by_status = {}
        for alert in self.alerts_db["alerts"].values():
            status = alert.get("status", "unknown")
            if status not in alerts_by_status:
                alerts_by_status[status] = []
            alerts_by_status[status].append(alert)

        # Display each group
        for status in ["new", "investigating", "escalated", "remediated", "closed"]:
            if status not in alerts_by_status:
                continue

            alerts = alerts_by_status[status]
            details.append(f"\n{status.upper()} ALERTS ({len(alerts)}):")
            details.append("-" * 70)

            for alert in sorted(alerts, key=lambda x: -x.get("normalized_severity", 0)):
                details.append(f"\n  ID: {alert['id']}")
                details.append(f"  Source: {alert.get('source', 'N/A'):15} | Severity: {alert.get('severity', 'N/A')}")
                details.append(f"  Title: {alert.get('title', 'N/A')[:60]}")
                if alert.get("is_false_positive"):
                    details.append("  ⚠️  Potential False Positive")
                if status == "remediated":
                    details.append(f"  Remediated: {alert.get('remediated_at', 'N/A')}")
                if status == "escalated":
                    details.append(f"  Escalated: {alert.get('escalated_at', 'N/A')}")
                    details.append(f"  Reason: {alert.get('escalation_reason', 'N/A')}")

        details.append("\n" + "=" * 70)

        return "\n".join(details)

    def export_json(self, output_path: Path):
        """
        Export summary to JSON format.

        Args:
            output_path: Output file path
        """
        stats = self.calculate_statistics()

        report = {
            "metadata": {
                "generated": datetime.utcnow().isoformat(),
                "report_type": "security_alert_summary",
                "version": "1.0",
            },
            "statistics": stats,
            "alerts_database": self.alerts_db,
            "remediation_log": self.remediation_log,
        }

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, default=str)

        print(f"✅ JSON report saved to {output_path}")

    def export_html(self, output_path: Path):
        """
        Export summary to HTML format.

        Args:
            output_path: Output file path
        """
        stats = self.calculate_statistics()
        exec_summary = self.generate_executive_summary(stats)
        breakdown = self.generate_detailed_breakdown(stats)

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Security Alert Summary Report</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        .metric-card {{ display: inline-block; background: #ecf0f1; padding: 15px 25px; margin: 10px; border-radius: 5px; min-width: 200px; }}
        .metric-label {{ font-size: 12px; color: #7f8c8d; text-transform: uppercase; }}
        .metric-value {{ font-size: 32px; font-weight: bold; color: #2c3e50; }}
        .critical {{ background: #e74c3c; color: white; }}
        .high {{ background: #e67e22; color: white; }}
        .success {{ background: #27ae60; color: white; }}
        .warning {{ background: #f39c12; color: white; }}
        pre {{ background: #2c3e50; color: #ecf0f1; padding: 20px; border-radius: 5px; overflow-x: auto; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #3498db; color: white; }}
        .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #7f8c8d; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🛡️ Security Alert Summary Report</h1>
        <p><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}</p>

        <h2>📊 Key Metrics</h2>
        <div>
            <div class="metric-card">
                <div class="metric-label">Total Alerts</div>
                <div class="metric-value">{stats['total_alerts']}</div>
            </div>
            <div class="metric-card success">
                <div class="metric-label">Remediated</div>
                <div class="metric-value">{stats['remediated_count']}</div>
            </div>
            <div class="metric-card warning">
                <div class="metric-label">Escalated</div>
                <div class="metric-value">{stats['escalated_count']}</div>
            </div>
            <div class="metric-card critical">
                <div class="metric-label">Critical Open</div>
                <div class="metric-value">{stats['critical_severity_open']}</div>
            </div>
            <div class="metric-card high">
                <div class="metric-label">High Open</div>
                <div class="metric-value">{stats['high_severity_open']}</div>
            </div>
        </div>

        <h2>📋 Executive Summary</h2>
        <pre>{exec_summary}</pre>

        <h2>📈 Detailed Breakdown</h2>
        <pre>{breakdown}</pre>

        <div class="footer">
            <p>Report generated by Easy-Ai Security Alert Investigation System</p>
            <p>For questions or issues, contact your security team.</p>
        </div>
    </div>
</body>
</html>
"""

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)

        print(f"✅ HTML report saved to {output_path}")

    def export_excel(self, output_path: Path):
        """
        Export summary to Excel format.

        Args:
            output_path: Output file path
        """
        if not PANDAS_AVAILABLE:
            print("❌ pandas not available - cannot export to Excel")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("❌ openpyxl not available - cannot export to Excel")
            return

        stats = self.calculate_statistics()

        # Create Excel workbook
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary.append(["Security Alert Summary Report"])
        ws_summary.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Alerts", stats["total_alerts"]])
        ws_summary.append(["Remediated", stats["remediated_count"]])
        ws_summary.append(["Escalated", stats["escalated_count"]])
        ws_summary.append(["Closed", stats["closed_count"]])
        ws_summary.append(["Pending", stats["pending_count"]])
        ws_summary.append(["Critical Open", stats["critical_severity_open"]])
        ws_summary.append(["High Open", stats["high_severity_open"]])
        ws_summary.append(["Remediation Rate (%)", stats["remediation_rate"]])
        ws_summary.append(["Closure Rate (%)", stats["closure_rate"]])

        # Format header
        for cell in ws_summary[1]:
            cell.font = Font(bold=True, size=14)
        for cell in ws_summary[4]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")

        # Alerts sheet
        ws_alerts = wb.create_sheet("Alerts")
        ws_alerts.append(["ID", "Source", "Severity", "Status", "Title", "Created", "Last Seen"])

        for alert in self.alerts_db["alerts"].values():
            ws_alerts.append(
                [
                    alert.get("id", ""),
                    alert.get("source", ""),
                    alert.get("severity", ""),
                    alert.get("status", ""),
                    alert.get("title", "")[:50],
                    alert.get("created", ""),
                    alert.get("last_seen", ""),
                ]
            )

        # Format alerts header
        for cell in ws_alerts[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        print(f"✅ Excel report saved to {output_path}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Generate comprehensive security alert summary reports",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        "--format",
        type=str,
        choices=["json", "html", "excel", "all"],
        default="all",
        help="Output format (default: all)",
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output/reports/security"),
        help="Output directory (default: output/reports/security)",
    )

    parser.add_argument(
        "--alerts-db",
        type=Path,
        default=Path("data/security/alerts.json"),
        help="Path to alerts database (default: data/security/alerts.json)",
    )

    parser.add_argument(
        "--remediation-log",
        type=Path,
        default=Path("output/reports/security/remediation_log.json"),
        help="Path to remediation log (default: output/reports/security/remediation_log.json)",
    )

    args = parser.parse_args()

    # Initialize generator
    generator = AlertSummaryGenerator(
        alerts_db_path=args.alerts_db,
        remediation_log_path=args.remediation_log,
    )

    # Calculate and display statistics
    stats = generator.calculate_statistics()
    print(generator.generate_executive_summary(stats))
    print(generator.generate_detailed_breakdown(stats))

    # Export to requested formats
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if args.format in ["json", "all"]:
        generator.export_json(args.output_dir / f"alert_summary_{timestamp}.json")

    if args.format in ["html", "all"]:
        generator.export_html(args.output_dir / f"alert_summary_{timestamp}.html")

    if args.format in ["excel", "all"]:
        generator.export_excel(args.output_dir / f"alert_summary_{timestamp}.xlsx")


if __name__ == "__main__":
    main()
